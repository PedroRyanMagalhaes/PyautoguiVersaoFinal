from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import pandas as pd
import pyautogui as pa
import time
from openpyxl.styles import PatternFill
import re
from openpyxl import load_workbook
import random


#Verde para pagos
fill_verde = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid") 
# Amarelo para atrasados ou em aberto
fill_amarelo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
#Vermelho para Cancelado/Linha não localizada/Saldo
fill_vermelho = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
# Laranja para não encontrar o mes
fill_laranja = PatternFill(start_color="FFFFA500", end_color="FFFFA500", fill_type="solid")
#roxo
fill_roxo = PatternFill(start_color="800080", end_color="800080", fill_type="solid")


#FUNCTIONS
def verificar_faturamento(driver):
    try:
        # Define o tempo máximo de espera
        wait = WebDriverWait(driver, 30)  # 10 segundos de timeout

        # Aguarda até que o elemento com o texto 'FATURAMENTO' esteja presente
        elemento_faturamento = wait.until(EC.presence_of_element_located((By.XPATH, "//span[contains(text(), 'FATURAMENTO')]")))

        print("Elemento 'FATURAMENTO' encontrado.")
        return True
    except Exception as e:
        tirar_print(driver,f"{linha_excel}erro.faturamento.png")
        print(f"Erro ao verificar 'FATURAMENTO': {e}")
        return False

def clicar_trespontos(driver, indice=2):
    try:
        # Espera até que o botão esteja clicável
        WebDriverWait(driver, 30).until(
            EC.element_to_be_clickable((By.XPATH, f"(//*[@aria-label='Opções'])[{indice}]"))
        )
        # Localiza e clica no botão
        botao_trespontos = driver.find_element(By.XPATH, f"(//*[@aria-label='Opções'])[{indice}]")
        botao_trespontos.click()
        return True  # Retorna True se clicou com sucesso
    except Exception as e:
        tirar_print(driver,f"{linha_excel}erro.trespontos1.png")

        print(f"Erro ao clicar nos três pontos: {e}")
        return False  # Retorna False se algo deu errado

def clicar_faturas(driver, indice=2):
    try:
        # Espera até que os botões estejam visíveis na página
        WebDriverWait(driver, 10).until(
            EC.visibility_of_all_elements_located((By.CSS_SELECTOR, 'button.MuiIconButton-root svg'))
        )
        
        # Encontra todos os botões de svg na página
        botoes_faturas = driver.find_elements(By.CSS_SELECTOR, 'button.MuiIconButton-root svg')
        
        # Clica no botão desejado com base no índice
        botoes_faturas[indice].click()
        return True  # Retorna True se clicou com sucesso
    except Exception as e:
        tirar_print(driver,f"{linha_excel}erro.faturas.png")
        print(f"Erro ao clicar no botão de faturas: {e}")
        return False  # Retorna False se algo deu errado

def mes1(driver, mes_esperado):
    try:
        # Localiza o elemento que contém a data usando o seletor fornecido
        elemento_mes = driver.find_element(By.CSS_SELECTOR, '#_value')
        data_texto = elemento_mes.text.strip()
        
        # Extrai o mês do formato esperado (dd/mm/aaaa)
        # Usamos regex para capturar somente o mês
        match = re.search(r'\d{2}/(\d{2})/\d{4}', data_texto)
        if match:
            mes_atual = match.group(1)  # Captura apenas o mês
            if mes_esperado == mes_atual:
                print(f"Mês encontrado no mes1: {mes_atual}")
                return True  # Se o mês esperado for encontrado
            else:
                print(f"Mês esperado: ({mes_esperado}) não corresponde ao mes1 ({mes_atual})")
                return False  # Se o mês não for encontrado
        else:
            print("Formato de data inválido ou mês não encontrado.")
            return False
    except Exception as e:
        tirar_print(driver,f"{linha_excel}erro.mes1.png")
        print(f"Erro ao procurar o mês no mes1: {e}")
        return True  # Retorna False se ocorrer algum erro


def mes2(driver, mes_esperado):
    try:
        # Localiza o quinto elemento com o seletor fornecido
        elemento_mes = driver.find_elements(By.CSS_SELECTOR, '#_value')[4]
        data_texto = elemento_mes.text.strip()

        #print(f"Mês encontrado mes2: {data_texto}")

        # Extrai o mês do formato esperado (dd/mm/aaaa)
        match = re.search(r'\d{2}/(\d{2})/\d{4}', data_texto)
        if match:
            mes_atual = match.group(1)  # Captura apenas o mês
            if mes_esperado == mes_atual:
                print(f"Mês encontrado no mes2: {mes_atual}")
                return True  # Se o mês esperado for encontrado
            else:
                print(f"Mês esperado ({mes_esperado}) não corresponde ao atual ({mes_atual})")
                return False  # Se o mês não for encontrado
        else:
            print("Formato de data inválido ou mês não encontrado.")
            return False
    except Exception as e:
        tirar_print(driver,f"{linha_excel}erro.mes2.png")
        print(f"Erro ao procurar o mês no mes2: {e}")
        return True  # Retorna False se ocorrer algum erro

def mes3(driver, mes_esperado):
    try:
        # Localiza o quinto elemento com o seletor fornecido
        elemento_mes = driver.find_elements(By.CSS_SELECTOR, '#_value')[8]
        data_texto = elemento_mes.text.strip()

        #print(f"Mês encontrado mes2: {data_texto}")

        # Extrai o mês do formato esperado (dd/mm/aaaa)
        match = re.search(r'\d{2}/(\d{2})/\d{4}', data_texto)
        if match:
            mes_atual = match.group(1)  # Captura apenas o mês
            if mes_esperado == mes_atual:
                print(f"Mês encontrado no mes3: {mes_atual}")
                return True  # Se o mês esperado for encontrado
            else:
                print(f"Mês esperado ({mes_esperado}) não corresponde ao atual ({mes_atual})")
                return False  # Se o mês não for encontrado
        else:
            print("Formato de data inválido ou mês não encontrado.")
            return False
    except Exception as e:
        tirar_print(driver,f"{linha_excel}erro.mes3.png")
        print(f"Erro ao procurar o mês no mes3: {e}")
        return True  # Retorna False se ocorrer algum erro



def status1(driver):
    try:
        # Captura todos os elementos com o seletor '#_value'
        elementos_status = driver.find_elements(By.CSS_SELECTOR, '#_value')
        
        # Obtém o terceiro elemento
        elemento_status = elementos_status[2]
        
        # Mostra o texto encontrado no terceiro elemento
        texto = elemento_status.text.strip().upper()  # Padroniza para maiúsculas
        #print(f"Texto encontrado no terceiro elemento: {texto}")
        
        # Verifica se o texto corresponde a um dos status esperados
        if texto in ['PAGA', 'ATRASADA', 'EM ABERTO']:
            print(f"Status encontrado no status1: {texto}")
            return texto  # Retorna o status encontrado
        
        # Caso o texto não corresponda a nenhum dos status esperados
        print(f"Texto encontrado no terceiro elemento não corresponde a um status válido: {texto}")
        return None
    except Exception as e:
        tirar_print(driver,f"{linha_excel}erro.status1.png")
        print(f"Erro ao verificar o status no status1: {str(e)}")
        return None

def status2(driver):
    try:
        # Captura todos os elementos com o seletor '#_value'
        elementos_status = driver.find_elements(By.CSS_SELECTOR, '#_value')
        
        # Obtém o terceiro elemento
        elemento_status = elementos_status[6]
        
        # Mostra o texto encontrado no terceiro elemento
        texto = elemento_status.text.strip().upper()  # Padroniza para maiúsculas
        #print(f"Texto encontrado no status 2: {texto}")
        
        # Verifica se o texto corresponde a um dos status esperados
        if texto in ['PAGA', 'ATRASADA', 'EM ABERTO']:
            print(f"Status encontrado no status2: {texto}")
            return texto  # Retorna o status encontrado
        
        # Caso o texto não corresponda a nenhum dos status esperados
        print(f"Texto encontrado no terceiro elemento não corresponde a um status válido: {texto}")
        return None
    except Exception as e:
        tirar_print(driver,f"{linha_excel}erro.status2.png")
        print(f"Erro ao verificar o status no status1: {str(e)}")
        return None

def status3(driver):
    try:
        # Captura todos os elementos com o seletor '#_value'
        elementos_status = driver.find_elements(By.CSS_SELECTOR, '#_value')
        
        # Obtém o terceiro elemento
        elemento_status = elementos_status[10]
        
        # Mostra o texto encontrado no terceiro elemento
        texto = elemento_status.text.strip().upper()  # Padroniza para maiúsculas
        #print(f"Texto encontrado no status 2: {texto}")
        
        # Verifica se o texto corresponde a um dos status esperados
        if texto in ['PAGA', 'ATRASADA', 'EM ABERTO']:
            print(f"Status encontrado no status3: {texto}")
            return texto  # Retorna o status encontrado
        
        # Caso o texto não corresponda a nenhum dos status esperados
        print(f"Texto encontrado no terceiro elemento não corresponde a um status válido: {texto}")
        return None
    except Exception as e:
        tirar_print(driver,f"{linha_excel}erro.status3.png")
        print(f"Erro ao verificar o status no status1: {str(e)}")
        return None

def verificar_cancelado(driver):
    try:
        # Espera até que o elemento seja visível
        elemento_cancelado = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="resultado-pesquisa"]/div[2]/div/div/div/div/div/div[5]/span[2]'))
        )
        
        # Verifica se o texto 'CANCELADO' está presente
        if "CANCELADO" in elemento_cancelado.text:

            print("Status: CANCELADO")
            return True
        else:
            print("Status não é CANCELADO")
            return False
    except Exception as e:
        tirar_print(driver,f"{linha_excel}erro.cancelado.png")
        print(f"Erro ao procurar o status CANCELADO: {e}")
        return False

def tres_pontos(driver):
    wait = WebDriverWait(driver, 30)  # 10 segundos de timeout

    try:
        # Espera até que o elemento esteja clicável
        elemento1 = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, '[data-src="/SVCv2/static/media/card_acelerator.7e6bc4b5.svg"]')))
        elemento1.click()
        return True
        
    except Exception as e:
        tirar_print(driver,f"{linha_excel}erro.trespontos2.png")
        print(f"Deu erro ao clicar nos três pontos: {e}")
        return False
    
def tirar_print(driver, nome_arquivo="erro_screenshot.png"):
    # Tira a captura de tela e salva no arquivo especificado
    driver.save_screenshot(nome_arquivo)
    print(f"Captura de tela salva em {nome_arquivo}")

def verificar_linha_nao_localizada(driver):
    try:
        # Espera até que o elemento seja visível na página
        elemento = WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.CSS_SELECTOR, '#layout > div > div.sc-VigVT.ldGKLs > h1'))
        )
        
        # Converte o texto do elemento para minúsculas e verifica
        if "linha não localizada" in elemento.text.lower():
            print("Texto 'Linha não localizada' detectado.")
            return True
        else:
            print("Texto esperado não encontrado no elemento.")
            return False
            
    except Exception as e:
        print(f"Erro ao verificar 'Linha não localizada': {e}")
        return False


def verificar_fatura_inexistente(driver):
    try:
        elemento = driver.find_element("xpath", '//*[@id="pagamento-body"]/div/div[1]/div[3]/div/div[2]/div/div/div[2]/p')
        if "Não existem faturas para este período" in elemento.text:
            return True
        return False
    except Exception as e:  # Captura qualquer erro e exibe
        print(f"Erro ao verificar fatura inexistente: {e}")
        return False

# Configurar o caminho do perfil do Chrome
options = webdriver.ChromeOptions()
#options.add_argument("user-data-dir=C:/Users/pedro/AppData/Local/Google/Chrome/User Data")  # Caminho da pasta User Data
#options.add_argument("profile-directory=Default")  # Usar o perfil "Default"

service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=service, options=options)

linha_inicial = 652 #Linha que começa a automação
linha_final = 952 #linha que finaliza automação 

# Acessar o site
driver.get("https://vendasapp.claro.com.br/SVCv2/posicionamento/resultado-pesquisa")

print("Insira o código de autenticação no site e clique em 'Entrar'. Depois, pressione Enter aqui no terminal para continuar.")
input("Pressione Enter quando estiver pronto...")

NomeDaPlanilha = ("janeiro.xlsx")

# Ler a planilha com pandas
mes_esperado = "01"
mes_seguinte = "02"
planilha = pd.read_excel(NomeDaPlanilha)
wb = load_workbook(NomeDaPlanilha)
ws = wb.active

wait = WebDriverWait(driver, 30) 

inico = time.time()

for linha_excel in range(linha_inicial, linha_final + 1):
    telefone = ws.cell(row=linha_excel, column=5).value
    
    print(f"Processando telefone: {telefone} da linha {linha_excel}")

    time.sleep(random.uniform(1, 2))  # Pausa entre 1 e 3 segundos
    campo_pesquisa = driver.find_element(By.ID, "filtro")
    campo_pesquisa.clear()  # Limpar o campo de pesquisa antes de inserir o próximo número
    campo_pesquisa.send_keys(str(telefone))
    campo_pesquisa.submit()

    
    if tres_pontos(driver):
        print("Cliquei nos tres pontos")
    elif verificar_cancelado(driver):
        # Pinta a linha de vermelho se o status for "CANCELADO"
            for col in range(1, ws.max_column + 1):
                ws.cell(row=linha_excel, column=col).fill = fill_vermelho  # Pinta toda a linha de vermelho
            print("Status CANCELADO - Linha pintada de vermelho")
            continue  # Volta para o começo do for e passa para a próxima linha
    else:
        if verificar_linha_nao_localizada(driver):
        # Pinta a linha de vermelho se o status for "CANCELADO"
            for col in range(1, ws.max_column + 1):
                ws.cell(row=linha_excel, column=col).fill = fill_vermelho  # Pinta toda a linha de vermelho
            print("Linha não localizada - Linha pintada de vermelho")
            pa.hotkey("alt","left")
            time.sleep(random.uniform(1,2))  # Pausa entre 1 e 3 segundos
            continue  # Volta para o começo do for e passa para a próxima linha

    

    elemento2 = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, '[data-src="/SVCv2/static/media/action_detalhe.f8a2c81b.svg"]')))
    elemento2.click()

    time.sleep(random.uniform(1, 2))  # Pausa entre 1 e 3 segundos

    if verificar_faturamento(driver):
        time.sleep(random.uniform(1, 2))  # Pausa entre 1 e 3 segundos
        if clicar_trespontos(driver):
            time.sleep(random.uniform(1, 2))  # Pausa entre 1 e 3 segundos
            print("Cliquei agora no segundo três pontos")
            if clicar_faturas(driver):
                time.sleep(random.uniform(1,2))  # Pausa entre 1 e 3 segundos
                print("Cliquei em faturas")
    else:
        for col in range(1, ws.max_column + 1):
                ws.cell(row=linha_excel, column=col).fill = fill_vermelho  # Pinta toda a linha de vermelho
        pa.hotkey("alt","left")
        time.sleep(random.uniform(1,2))  # Pausa entre 1 e 3 segundos
        print("Não achei faturamento")
        continue  
        
    # Verificando o mês
    status = None
    mes_esperado_encontrado = False
    mes_seguinte_encontrado = False

    try:
    # Verifica se o mês esperado foi encontrado
        if mes1(driver, mes_esperado):
          status = status1(driver)
          mes_esperado_encontrado = True
        elif mes2(driver, mes_esperado):
            status = status2(driver)
            mes_esperado_encontrado = True
        elif mes3(driver, mes_esperado):
            status = status3(driver)
            mes_esperado_encontrado = True
    except IndexError as e:
        print(f"Erro ao procurar o mês esperado: {e}")
        status = None
        mes_esperado_encontrado = False

# Se o mês esperado não foi encontrado, tenta encontrar o mês seguinte
    if mes_esperado_encontrado == False:
        try:
            if mes1(driver, mes_seguinte) or mes2(driver, mes_seguinte) or mes3(driver, mes_seguinte):
                print (f"Mes seguinte = {mes_seguinte}")
                mes_seguinte_encontrado = True
                status = "ROXO"  # Alterando o status para "ROXO" se o mês seguinte for encontrado
        except IndexError as e:
            print(f"Erro ao procurar o mês seguinte: {e}")
            mes_seguinte_encontrado = False

    if not mes_esperado_encontrado and not mes_seguinte_encontrado:
        verificar_fatura_inexistente(driver)
        print ("Fatura inexistente ")
        status = "ROXO"


# Verifica o que fazer baseado no status
    if status == "ROXO":
        print("Mês seguinte encontrado, mas mês esperado não")
        for col in range(1, ws.max_column + 1):
            ws.cell(row=linha_excel, column=col).fill = fill_roxo
    elif mes_esperado_encontrado:
        print("Mês esperado encontrado")
        if status == "PAGA":
            print("Status: Paga")
            for col in range(1, ws.max_column + 1):
                ws.cell(row=linha_excel, column=col).fill = fill_verde
        elif status in ["ATRASADA", "EM ABERTO"]:
            print(f"Status: {status}")
            for col in range(1, ws.max_column + 1):
                ws.cell(row=linha_excel, column=col).fill = fill_amarelo
        else:
            print("Status não encontrado")
            for col in range(1, ws.max_column + 1):
                ws.cell(row=linha_excel, column=col).fill = fill_laranja
    else:
        print("Nenhum mês encontrado")
        for col in range(1, ws.max_column + 1):
            ws.cell(row=linha_excel, column=col).fill = fill_laranja
   


    # Volta para a página anterior
    pa.hotkey("alt", "left")
    pa.hotkey("alt", "left")
    time.sleep(random.uniform(1, 2))
    wb.save(NomeDaPlanilha)

fim = time.time()
totaltime = fim - inico
totalminute = (totaltime / 60)
totalconferido = linha_final - linha_inicial

print(f"Tempo total foi {totaltime:.2f} segundos e {totalminute:.2f} minutos para conferir {totalconferido}")
driver.quit()



