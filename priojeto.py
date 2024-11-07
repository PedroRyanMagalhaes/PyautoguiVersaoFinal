

import openpyxl
import datetime
import pyautogui as pa  
import time
import pyperclip
import pytesseract
import cv2
from openpyxl.styles import PatternFill
import os


# Carrega a planilha
wb = openpyxl.load_workbook('OUT.xlsx')
sheet = wb.active


def esperar_carregamento(imagem_carregando, confiança=0.8, pasta_salvamento='capturas', max_tentativas=100):

    # Cria a pasta para salvar as capturas, se não existir

    tentativas = 0

    while tentativas < max_tentativas:
        try:
            posicao = pa.locateOnScreen(imagem_carregando, confidence=confiança)


            if posicao is None:
                break
            else:
                time.sleep(0.2)  # Verifica a cada 100 milissegundos
                tentativas += 1  # Incrementa o contador de tentativas
        except Exception as e:
            print(f"Ocorreu um erro: {e}")
            break  # Em caso de erro, saia do loop
    
    if tentativas >= max_tentativas:
        print("Número máximo de tentativas alcançado. Saindo do loop.")





def criarPastaParaLinha(linha):
    pasta_ = f'prints{linha}'
    if not os.path.exists(pasta_):
        os.makedirs(pasta_)
    return pasta_

def verificarStatus(linha):
    status = verificarStatusPagamento(linha) or verificarStatusPagamentoAlternativo(linha)
    return status

def obterTelefone(linha):   
    return sheet[f'E{linha}'].value

def pintarDeVerde(linha):
    verde = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
    for cell in sheet[linha]:
        cell.fill = verde
    

def pintarDeVermelho(linha):
    vermelho = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    for cell in sheet[linha]:
        cell.fill = vermelho
    

def pintarDeAmarelo(linha):
    amarelo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for cell in sheet[linha]:
        cell.fill = amarelo

# Função para verificar o mês
def verificaMes(linha, mesEsperado):
    try:
        pasta_ = criarPastaParaLinha(linha)
        screenshot_mes = os.path.join(pasta_, f'{linha}_printMes.png')
        screenshot = pa.screenshot(region=(524, 705, 100, 60))  # Ajuste a região conforme necessário
        screenshot.save(screenshot_mes)

        img = cv2.imread(screenshot_mes)

        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        blurred = cv2.GaussianBlur(gray, (3, 3), 0)
        _, thresh = cv2.threshold(blurred, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)

        extracted_text = pytesseract.image_to_string(thresh, config='--psm 6 -c tessedit_char_whitelist=0123456789/')
        print(f"Mes: {linha}: {extracted_text}")

        return mesEsperado in extracted_text

    except Exception as e:
        print(f"Ocorreu um erro ao verificar o mês na linha {linha}: {str(e)}")
        return False


# Função para verificar o mês em uma nova coordenada
def verificarMesAlternativo(linha, mesEsperado):
    try:
        pasta_ = criarPastaParaLinha(linha)
        screenshot_mes = os.path.join(pasta_, f'{linha}_printMesAlt.png')
        screenshot = pa.screenshot(region=(525,781, 100, 40))  # Ajuste a região conforme necessário
        screenshot.save(screenshot_mes)

        img = cv2.imread(screenshot_mes)

        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        blurred = cv2.GaussianBlur(gray, (3, 3), 0)
        _, thresh = cv2.threshold(blurred, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)

        extracted_text = pytesseract.image_to_string(thresh, config='--psm 6 -c tessedit_char_whitelist=0123456789/')
        print(f"Mes Alt {linha}: {extracted_text}")

        return mesEsperado in extracted_text

    except Exception as e:
        print(f"Ocorreu um erro ao verificar o mês alternativo na linha {linha}: {str(e)}")
        return False
# Função para verificar se está "pago" ou "atrasada"
def verificarStatusPagamento(linha):
    try:
        pasta_ = criarPastaParaLinha(linha)
        screenshot_status = os.path.join(pasta_, f'{linha}_printPagamento.png')
        screenshot = pa.screenshot(region=(1024,703, 100, 40))  # Ajuste a região conforme necessário
        screenshot.save(screenshot_status)

        img = cv2.imread(screenshot_status)

        # Aumentar a resolução da imagem para melhorar a leitura
        img = cv2.resize(img, None, fx=2, fy=2, interpolation=cv2.INTER_CUBIC)
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        blurred = cv2.GaussianBlur(gray, (1, 1), 0)
        _, thresh = cv2.threshold(blurred, 150, 255, cv2.THRESH_BINARY_INV)

        extracted_text = pytesseract.image_to_string(thresh, config='--psm 7')
        print(f"Texto extraído para o pagamento na linha {linha}: {extracted_text}")

        if "paga" in extracted_text.lower():
            print(f"Status 'paga' encontrado na linha {linha}.")
            return "paga"
        elif "atrasada" in extracted_text.lower():
            print(f"Status 'atrasada' encontrado na linha {linha}.")
            return "atrasada"
        else:
            print(f"Status desconhecido na linha {linha}: {extracted_text}")
            return None

    except Exception as e:
        print(f"Ocorreu um erro ao verificar o status de pagamento na linha {linha}: {str(e)}")
        return "erro"


# Função para verificar o status de pagamento em uma nova coordenada
def verificarStatusPagamentoAlternativo(linha):
    try:
        pasta_ = criarPastaParaLinha(linha)
        screenshot_status = os.path.join(pasta_, f'{linha}_printPagamentoAlt.png')
        screenshot = pa.screenshot(region=(1027,783, 130, 50))  # Ajuste a região conforme necessário
        screenshot.save(screenshot_status)

        img = cv2.imread(screenshot_status)

        # Aumentar a resolução da imagem para melhorar a leitura
        img = cv2.resize(img, None, fx=2, fy=2, interpolation=cv2.INTER_CUBIC)
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        blurred = cv2.GaussianBlur(gray, (1, 1), 0)
        _, thresh = cv2.threshold(blurred, 150, 255, cv2.THRESH_BINARY_INV)

        extracted_text = pytesseract.image_to_string(thresh, config='--psm 7')
        print(f"Texto extraído para o pagamento na linha Alternativo {linha}: {extracted_text}")

        if "paga" in extracted_text.lower():
            print(f"Status 'paga' encontrado na linha {linha}.")
            return "paga"
        elif "atrasada" in extracted_text.lower():
            print(f"Status 'atrasada' encontrado na linha {linha}.")
            return "atrasada"
        else:
            print(f"pagamento desconhecido na linha alternativa {linha}: {extracted_text}")
            return None

    except Exception as e:
        print(f"Ocorreu um erro ao verificar o status de pagamento na linha {linha}: {str(e)}")
        return "erro"


def verificarFaturamento(linha):
    try:
        pasta_ = criarPastaParaLinha(linha)
        screenshot_faturamento = os.path.join(pasta_, f'{linha}_printFaturamento.png')
        screenshot = pa.screenshot(region=(491,474,400,100))  # Ajuste essa coordenada para a região correta
        screenshot.save(screenshot_faturamento)

        img = cv2.imread(screenshot_faturamento)

        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        _, thresh = cv2.threshold(gray, 150, 255, cv2.THRESH_BINARY_INV)

        extracted_text = pytesseract.image_to_string(thresh)
        print(f"Texto extraído para faturamento na linha {linha}: {extracted_text}")

        if "faturamento" in extracted_text.lower():
            print(f"'Faturamento' encontrado na linha {linha}.")
            return True
        else:
            print(f"'Faturamento' não encontrado na linha {linha}.")
            return False

    except Exception as e:
        print(f"Ocorreu um erro ao verificar o 'faturamento' na linha {linha}: {str(e)}")
        return False

# Função para verificar "faturamento" em uma coordenada alternativa
def verificarFaturamentoAlternativo(linha):
    try:
        pasta_ = criarPastaParaLinha(linha)
        screenshot_faturamento_alt = os.path.join(pasta_, f'{linha}_printFaturamentoAlt.png')
        screenshot = pa.screenshot(region=(503,623, 500, 250))  # Ajuste essa coordenada para a nova região
        screenshot.save(screenshot_faturamento_alt)

        img = cv2.imread(screenshot_faturamento_alt)

        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        _, thresh = cv2.threshold(gray, 150, 255, cv2.THRESH_BINARY_INV)

        extracted_text = pytesseract.image_to_string(thresh)
        print(f"Texto extraído para faturamento na linha {linha} (alternativo): {extracted_text}")

        if "faturamento" in extracted_text.lower():
            print(f"'Faturamento' encontrado na linha Alt {linha}.")
            return True
        else:
            return False

    except Exception as e:
        print(f"Ocorreu um erro ao verificar o 'faturamento' alternativo na linha {linha}: {str(e)}")
        return False
    
def verificarSaldo(linha):
    try:
        pasta_ = criarPastaParaLinha(linha)
        screenshot_saldo = os.path.join(pasta_, f'{linha}_printSaldo.png')
        screenshot = pa.screenshot(region=(505,474,400,100))  # Ajuste essa coordenada para a região correta
        screenshot.save(screenshot_saldo)

        img = cv2.imread(screenshot_saldo)

        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        _, thresh = cv2.threshold(gray, 150, 255, cv2.THRESH_BINARY_INV)

        extracted_text = pytesseract.image_to_string(thresh)
        print(f"Texto extraído para saldo na linha {linha}: {extracted_text}")

        if "saldo" in extracted_text.lower():
            print(f"'Saldo' encontrado na linha {linha}.")
            return True
        else:
            return False

    except Exception as e:
        print(f"Ocorreu um erro ao verificar o 'saldo' na linha {linha}: {str(e)}")
        return False

    
contador = 1

def clicarTresPontos(imagem_tres_pontos, regiao=None):
    # Aguarda um momento para garantir que a tela esteja pronta
    time.sleep(1)

    # Captura a tela ou a região especificada
    pasta_ = criarPastaParaLinha(linha)
    screenshot = pa.screenshot(region=regiao)
    nome_arquivo = os.path.join(pasta_, f"{linha}trespontos.png")  # Salva a captura para debug
    screenshot.save(nome_arquivo)
    print(f"Captura de tela salva como: {nome_arquivo}")

    posicao = None

    # Tenta localizar a imagem na tela
    try:
        posicao = pa.locateOnScreen(imagem_tres_pontos, confidence=0.5, region=regiao)
        if posicao:
            print(f"Imagem dos três pontos localizada na posição: {posicao}")
        else:
            print("Imagem dos três pontos não encontrada na tela.")
    except Exception as e:
        print(f"Erro ao tentar localizar a imagem: {e}")
        return False  # Retorna False se ocorrer qualquer erro

    if posicao is not None:
        # Clica no centro da imagem encontrada
        pa.click(pa.center(posicao))
        print("Clicou nos três pontinhos.")
        return True
    else:
        return False


    
def clicarfaturas(imagem_faturas, regiao=None):
    
    
    # Aguarda um momento para garantir que a tela esteja pronta
    time.sleep(1)

    pasta_ = criarPastaParaLinha(linha)

    # Captura a tela ou a região especificada
    screenshot = pa.screenshot(region=regiao)

    nome_arquivo = os.path.join(pasta_, f"{linha}faturas.png")

    screenshot.save(nome_arquivo)

    # Verifica se a imagem dos faturas está presente na captura
    posicao = pa.locateOnScreen(imagem_faturas, confidence=0.5, region=regiao)
    
    if posicao is not None:
        # Clica no centro da imagem encontrada
        pa.click(pa.center(posicao))
        print("Clicou em faturas.")
        return True
    else:
        print("Imagem dos faturas não encontrada.")
        return False

def verificarLinhaNaoLocalizada(linha):
        pasta_ = criarPastaParaLinha(linha)
        screenshot_linhanaolocalizada = os.path.join(pasta_, f'{linha}_LinhaNaoLocalizada.png')
        screenshot = pa.screenshot(region=(827,430,400,400))  # Ajuste essa coordenada para a região correta
        screenshot.save(screenshot_linhanaolocalizada)

        img = cv2.imread(screenshot_linhanaolocalizada)

        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        _, thresh = cv2.threshold(gray, 150, 255, cv2.THRESH_BINARY_INV)

        extracted_text = pytesseract.image_to_string(thresh)
        print(f"Texto extraído na linha {linha}: {extracted_text}")

        if ("linha não localizada" in extracted_text.lower() or 
            "localizada" in extracted_text.lower() or 
            "venda avulsa" in extracted_text.lower()):
             print(f"Essa linha é não localizada")
        else:
            return False

# Mês esperado
mesEsperado = "10"


comecoLinha = 2
finalLinha = 2

horario_inicial = datetime.datetime.now().strftime("%H:%M:%S")


# Muda a tela
pa.hotkey('alt', 'tab')
time.sleep(0.3)

with open('telefone.txt', 'w') as f:
    for linha in range(comecoLinha, finalLinha + 1):
        telefone = obterTelefone(linha)
        f.write(f"linha {linha}, Telefone {telefone}\n")



for linha in range(comecoLinha, finalLinha + 1):
    telefone = obterTelefone(linha)  
    pyperclip.copy(telefone)  


    # Clica no lugar para dar 'ctrl + a'
    
    esperar_carregamento('assets/PCcarregando.png',0.5)
    pa.click(690,294)
    pa.hotkey('ctrl', 'a')
    time.sleep(0.3)

    # Cola o telefone
    pa.hotkey('ctrl', 'v')
    pa.press('enter')
    time.sleep(0.3)
    

    esperar_carregamento('assets/PCcarregando.png',0.5)
    

    # Clicar nos 3 pontinhos
    time.sleep(0.5)
    resultado = clicarTresPontos(imagem_tres_pontos='assets/PCtrespontos.PNG', regiao=(670,328, 100, 100))
    
    
    if resultado:
        time.sleep(0.3)
    else:
    # Clica em uma posição fixa se o resultado for False
        pa.click(623, 516)
        print("Clicou na posição fixa (623, 516) para tentar acessar os três pontos.")
        resultado = clicarTresPontos(imagem_tres_pontos='assets/PCtrespontos.PNG', regiao=(670, 328, 100, 100))
    
        if resultado:
            print("Imagem dos três pontos localizada e clicada com sucesso.")
            time.sleep(0.3)
        else:
            print("Não foi possível localizar a imagem dos três pontos após o clique na posição fixa.")
            verificarLinhaNaoLocalizada(linha)
            pintarDeVermelho(linha)
            pa.hotkey('alt', 'left')
            time.sleep(0.5)
            print(f"Processamento da linha {linha} = linha não localizada.")
            time.sleep(0.5)
            continue

            
            

    # Clicar em detalhes 
    pa.click(546,380)
    time.sleep(0.5)
    esperar_carregamento('assets/carregando.jpg',0.5)
    time.sleep(2)

    
    if verificarSaldo(linha):
            pintarDeVermelho(linha)
            print(f'Saldo encontrado na linha {linha}. Linha pintada de vermeleho.')
            pa.hotkey('alt', 'left')  # Voltar com Alt + Left
            #time.sleep(0.5)
            esperar_carregamento('assets/PCcarregando.png',0.5)
            continue  # Volta para o início do for
    
        


  # Verificar "faturamento" após clicar em detalhes
    if verificarFaturamento(linha):
        imagem_tres_pontos = 'assets/imagemtrespontos.jpg'  # Substitua pelo caminho real da sua imagem
        regiao = (868,489, 200, 280)  # Substitua pelas coordenadas (x, y, largura, altura) região que deseja capturar
    
    
        if clicarTresPontos(imagem_tres_pontos, regiao):
            time.sleep(0.5)  # Aguardar um tempo após clicar nos três pontos

        # Clicar em faturas
        imagem_faturas = 'assets/imagemfaturas.jpg'  # Clicar em faturas
        regiao = (522,496, 200, 200)

        if clicarfaturas(imagem_faturas, regiao):
            time.sleep(0.5)
         
    else:
        pa.scroll(-500)  # Scroll para baixo
        time.sleep(1)

    # Se "faturamento" não for encontrado, verificar na coordenada alternativa
    if verificarFaturamentoAlternativo(linha):
        # Se encontrado na coordenada alternativa, realizar scroll e cliques adicionais

            imagem_tres_pontos = 'assets/imagemtrespontos.jpg'  # Clique em nova coordenada
            regiao = (868,631, 250, 280)

            if clicarTresPontos(imagem_tres_pontos, regiao):
                time.sleep(0.5)

            imagem_faturas = 'assets/imagemfaturas.jpg'  # Outro clique
            regiao = (505,652, 250, 380)

            if clicarfaturas(imagem_faturas, regiao):
                time.sleep(0.5)

            pa.scroll(500)  # Scroll para cima
            time.sleep(1)
        
        

    else:
        # Se "faturamento" não for encontrado em nenhuma das coordenadas, pular linha
            print(f"'Faturamento' não encontrado na linha alternativa {linha}.")
         # Pula para a próxima linha se nenhuma verificação passar

# Verificar o mês capturado na tela
    if verificaMes(linha,mesEsperado):
    # Verificar o status de pagamento
        status = verificarStatusPagamento(linha)
        print(f"Status para a linha {linha}: {status}")
        if status == "paga":
            pintarDeVerde(linha)
        elif status == "atrasada":
            pintarDeAmarelo(linha)
        else:
            pintarDeVermelho(linha)
    else:
    # Se o mês não for o esperado, tenta verificar em uma nova coordenada
        if verificarMesAlternativo(linha, mesEsperado):
        # Verificar o status de pagamento na coordenada alternativa
            status = verificarStatusPagamentoAlternativo(linha)
            print(f"Status alternativo para a linha {linha}: {status}")
            if status == "paga":
                pintarDeVerde(linha)
            elif status == "atrasada":
                pintarDeAmarelo(linha)
        else:
        # Se o mês ainda não for o esperado, pinta a linha de laranja
            pintarDeVermelho(linha)
            print ("Pintei de vermelho, verificar depois por favor")
        

# Voltar para a tela anterior duas vezes
    pa.hotkey('alt', 'left')
    time.sleep(0.5)
    pa.hotkey('alt', 'left')
    #time.sleep(7)
    #esperar_carregamento('assets/carregando.jpg',0.5)
    wb.save("out10.xlsx")

print (f"Começou às {horario_inicial}")
horario_final = datetime.datetime.now().strftime("%H:%M:%S")

print(f"Processo finalizado para todas as linhas às {horario_final}")